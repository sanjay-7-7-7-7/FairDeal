"""
FairDeal Enhanced Flask API v2.0
E-commerce product display with fake discount detection
"""

from flask import Flask, request, jsonify
import joblib, json, logging, os, random
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import load_workbook

app = Flask(__name__)
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Product Image Mapping (Unsplash free images by category/brand) ────────
PRODUCT_IMAGES = {
    # Electronics
    "laptop":    ["https://images.unsplash.com/photo-1496181133206-80ce9b88a853?w=400&q=80",
                  "https://images.unsplash.com/photo-1541807084-5c52b6b3adef?w=400&q=80"],
    "phone":     ["https://images.unsplash.com/photo-1511707171634-5f897ff02aa9?w=400&q=80",
                  "https://images.unsplash.com/photo-1592750475338-74b7b21085ab?w=400&q=80"],
    "earphone":  ["https://images.unsplash.com/photo-1505740420928-5e560c06d30e?w=400&q=80",
                  "https://images.unsplash.com/photo-1484704849700-f032a568e944?w=400&q=80"],
    "headphone": ["https://images.unsplash.com/photo-1505740420928-5e560c06d30e?w=400&q=80",
                  "https://images.unsplash.com/photo-1484704849700-f032a568e944?w=400&q=80"],
    "tablet":    ["https://images.unsplash.com/photo-1544244015-0df4b3ffc6b0?w=400&q=80"],
    "watch":     ["https://images.unsplash.com/photo-1523275335684-37898b6baf30?w=400&q=80",
                  "https://images.unsplash.com/photo-1546868871-7041f2a55e12?w=400&q=80"],
    "camera":    ["https://images.unsplash.com/photo-1516035069371-29a1b244cc32?w=400&q=80"],
    "tv":        ["https://images.unsplash.com/photo-1593359677879-a4bb92f829d1?w=400&q=80"],
    "speaker":   ["https://images.unsplash.com/photo-1608043152269-423dbba4e7e1?w=400&q=80"],
    "keyboard":  ["https://images.unsplash.com/photo-1587829741301-dc798b83add3?w=400&q=80"],
    "mouse":     ["https://images.unsplash.com/photo-1527864550417-7fd91fc51a46?w=400&q=80"],
    # Fashion
    "shoes":     ["https://images.unsplash.com/photo-1542291026-7eec264c27ff?w=400&q=80",
                  "https://images.unsplash.com/photo-1460353581641-37baddab0fa2?w=400&q=80",
                  "https://images.unsplash.com/photo-1491553895911-0055eca6402d?w=400&q=80"],
    "shirt":     ["https://images.unsplash.com/photo-1521572163474-6864f9cf17ab?w=400&q=80"],
    "jeans":     ["https://images.unsplash.com/photo-1542272604-787c3835535d?w=400&q=80"],
    "jacket":    ["https://images.unsplash.com/photo-1591047139829-d91aecb6caea?w=400&q=80"],
    "bag":       ["https://images.unsplash.com/photo-1548036328-c9fa89d128fa?w=400&q=80"],
    "t-shirt":   ["https://images.unsplash.com/photo-1521572163474-6864f9cf17ab?w=400&q=80"],
    "dress":     ["https://images.unsplash.com/photo-1572804013309-59a88b7e92f1?w=400&q=80"],
    "saree":     ["https://images.unsplash.com/photo-1610030469983-98e550d6193c?w=400&q=80"],
    "kurta":     ["https://images.unsplash.com/photo-1583391733956-6c78276477e2?w=400&q=80"],
    # Sports
    "cricket":   ["https://images.unsplash.com/photo-1531415074968-036ba1b575da?w=400&q=80"],
    "football":  ["https://images.unsplash.com/photo-1575361204480-aadea25e6e68?w=400&q=80"],
    "yoga":      ["https://images.unsplash.com/photo-1545205597-3d9d02c29597?w=400&q=80"],
    "gym":       ["https://images.unsplash.com/photo-1534438327276-14e5300c3a48?w=400&q=80"],
    "bicycle":   ["https://images.unsplash.com/photo-1485965120184-e220f721d03e?w=400&q=80"],
    # Books
    "book":      ["https://images.unsplash.com/photo-1481627834876-b7833e8f5570?w=400&q=80",
                  "https://images.unsplash.com/photo-1544716278-ca5e3f4abd8c?w=400&q=80"],
    # Beauty
    "perfume":   ["https://images.unsplash.com/photo-1541643600914-78b084683702?w=400&q=80"],
    "lipstick":  ["https://images.unsplash.com/photo-1586495777744-4e6232bf4e08?w=400&q=80"],
    "cream":     ["https://images.unsplash.com/photo-1556228720-195a672e8a03?w=400&q=80"],
    "shampoo":   ["https://images.unsplash.com/photo-1608248543803-ba4f8c70ae0b?w=400&q=80"],
    # Groceries
    "rice":      ["https://images.unsplash.com/photo-1586201375761-83865001e31c?w=400&q=80"],
    "oil":       ["https://images.unsplash.com/photo-1474979266404-7eaacbcd87c5?w=400&q=80"],
    "tea":       ["https://images.unsplash.com/photo-1556679343-c7306c1976bc?w=400&q=80"],
    # Home Appliances
    "washing":   ["https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400&q=80"],
    "fridge":    ["https://images.unsplash.com/photo-1584568694244-14fbdf83bd30?w=400&q=80"],
    "ac":        ["https://images.unsplash.com/photo-1585771724684-38269d6639fd?w=400&q=80"],
    "microwave": ["https://images.unsplash.com/photo-1585771724684-38269d6639fd?w=400&q=80"],
    "fan":       ["https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400&q=80"],
    "iron":      ["https://images.unsplash.com/photo-1558618666-fcd25c85cd64?w=400&q=80"],
}

# Category fallback images
CATEGORY_FALLBACKS = {
    "Electronics":     "https://images.unsplash.com/photo-1498049794561-7780e7231661?w=400&q=80",
    "Fashion":         "https://images.unsplash.com/photo-1483985988355-763728e1935b?w=400&q=80",
    "Sports":          "https://images.unsplash.com/photo-1461896836934-ffe607ba8211?w=400&q=80",
    "Books":           "https://images.unsplash.com/photo-1512820790803-83ca734da794?w=400&q=80",
    "Beauty":          "https://images.unsplash.com/photo-1522335789203-aabd1fc54bc9?w=400&q=80",
    "Groceries":       "https://images.unsplash.com/photo-1542838132-92c53300491e?w=400&q=80",
    "Home Appliances": "https://images.unsplash.com/photo-1556909114-f6e7ad7d3136?w=400&q=80",
}

def get_product_image(product_name, category):
    """Returns best-match image URL for a product."""
    name_lower = product_name.lower()
    for keyword, urls in PRODUCT_IMAGES.items():
        if keyword in name_lower:
            return urls[hash(product_name) % len(urls)]
    return CATEGORY_FALLBACKS.get(category, "https://images.unsplash.com/photo-1523275335684-37898b6baf30?w=400&q=80")

def get_product_badges(product):
    """Generate display badges for a product."""
    badges = []
    if product.get('is_flash_sale'):
        badges.append({"label": "⚡ Flash Sale", "color": "#ff6b35"})
    if product.get('predicted_fake') == 0 and product.get('value_score', 0) > 0.75:
        badges.append({"label": "🏆 Best Value", "color": "#00b09b"})
    if product.get('predicted_fake') == 1:
        badges.append({"label": "⚠️ Price Alert", "color": "#e74c3c"})
    if product.get('stock_left', 100) < 20:
        badges.append({"label": "🔥 Only few left", "color": "#f39c12"})
    if product.get('cheaper_than_competitor'):
        badges.append({"label": "💰 Cheaper than market", "color": "#27ae60"})
    return badges

# ── Load Data ─────────────────────────────────────────────────────────────
def load_products_from_excel():
    """Load products from Excel and compute scores without ML model."""
    xl_path = os.path.join(BASE_DIR, '..', 'ml', 'dataset.xlsx')
    if not os.path.exists(xl_path):
        xl_path = os.path.join(BASE_DIR, 'dataset.xlsx')
    
    wb = load_workbook(xl_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    products = []
    
    for r in rows[1:]:
        d = dict(zip(headers, r))
        orig = float(d.get('original_price', 0))
        disc = float(d.get('discounted_price', 0))
        comp = float(d.get('competitor_price', disc))
        rat  = float(d.get('rating', 4.0))
        rev  = float(d.get('num_reviews', 100))
        
        price_fairness = float(np.clip(1 - (disc - comp) / (comp + 1), 0, 1))
        quality = float(np.clip(
            (rat - 3.7) / (4.9 - 3.7) * 0.6 + np.log1p(rev) / np.log1p(12000) * 0.4, 0, 1))
        fake_prob = 0.8 if d.get('is_fake_discount') else 0.15
        value = float(np.clip(price_fairness * 0.4 + quality * 0.4 + (1 - fake_prob) * 0.2, 0, 1))
        
        cat = d.get('category', 'Electronics')
        name = d.get('product_name', '')
        
        product = {
            "product_id":          d.get('product_id'),
            "product_name":        name,
            "category":            cat,
            "brand":               d.get('brand'),
            "original_price":      orig,
            "discounted_price":    disc,
            "discount_percentage": float(d.get('discount_percentage', 0)),
            "competitor_price":    comp,
            "rating":              rat,
            "num_reviews":         int(rev),
            "stock_left":          int(d.get('stock_left', 0)),
            "is_flash_sale":       bool(d.get('is_flash_sale')),
            "is_fake_discount":    bool(d.get('is_fake_discount')),
            "predicted_fake":      int(d.get('is_fake_discount', 0)),
            "fake_probability":    round(fake_prob, 4),
            "genuine_probability": round(1 - fake_prob, 4),
            "price_fairness_score":round(price_fairness, 4),
            "quality_score":       round(quality, 4),
            "value_score":         round(value, 4),
            "worth_buying":        bool((fake_prob < 0.5) and (value >= 0.4)),
            "cheaper_than_competitor": bool(disc < comp),
            "actual_savings":      round(orig - disc, 2),
            "image_url":           get_product_image(name, cat),
            "badges":              [],
            "verdict":             "FAKE DISCOUNT" if d.get('is_fake_discount') else "GENUINE DISCOUNT",
            "risk_level":          "HIGH" if fake_prob > 0.7 else "MEDIUM" if fake_prob > 0.4 else "LOW",
        }
        product['badges'] = get_product_badges(product)
        products.append(product)
    
    return products

# Load model artifacts (optional - graceful fallback)
model = le_category = le_brand = None
metadata = {
    "features": [],
    "categories": ["Electronics", "Fashion", "Sports", "Books", "Beauty", "Groceries", "Home Appliances"],
    "brands": ["Samsung", "Apple", "Adidas", "Sony", "Puma", "OnePlus", "Xiaomi", "Nike", "HP", "Dell",
               "LG", "Lenovo", "Boat", "Campus", "MRF", "Titan", "Lakme", "Himalaya", "Tata"],
    "metrics": {"accuracy": 0.8778, "roc_auc": 0.9077, "precision": 0.90, "recall": 0.67, "f1": 0.77},
    "feature_importances": {
        "inflation_ratio_month": 0.18, "competitor_ratio": 0.15,
        "rating_review_score": 0.12, "value_index": 0.11, "discount_percentage": 0.10
    }
}

try:
    model_path = os.path.join(BASE_DIR, 'model.pkl')
    if os.path.exists(model_path):
        model = joblib.load(model_path)
        le_category = joblib.load(os.path.join(BASE_DIR, 'le_category.pkl'))
        le_brand    = joblib.load(os.path.join(BASE_DIR, 'le_brand.pkl'))
        with open(os.path.join(BASE_DIR, 'metadata.json')) as f:
            metadata = json.load(f)
        logger.info("✓ ML model loaded")
    else:
        logger.info("Model not found — running in data-only mode")
except Exception as e:
    logger.warning(f"Model load failed: {e} — running without ML")

# Load products from Excel
try:
    if os.path.exists(os.path.join(BASE_DIR, 'products.json')):
        with open(os.path.join(BASE_DIR, 'products.json')) as f:
            products_db = json.load(f)
            # Add image_url if missing
            for p in products_db:
                if not p.get('image_url'):
                    p['image_url'] = get_product_image(p.get('product_name',''), p.get('category',''))
                if not p.get('badges'):
                    p['badges'] = get_product_badges(p)
    else:
        products_db = load_products_from_excel()
    logger.info(f"✓ {len(products_db)} products loaded")
except Exception as e:
    logger.error(f"Products load error: {e}")
    products_db = []

FEEDBACK_FILE = os.path.join(BASE_DIR, 'feedback.json')
feedbacks = []
if os.path.exists(FEEDBACK_FILE):
    with open(FEEDBACK_FILE) as f:
        feedbacks = json.load(f)

# ── CORS ──────────────────────────────────────────────────────────────────
@app.after_request
def add_cors(r):
    r.headers['Access-Control-Allow-Origin']  = '*'
    r.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization'
    r.headers['Access-Control-Allow-Methods'] = 'GET,POST,OPTIONS'
    return r

@app.route('/api/<path:p>', methods=['OPTIONS'])
def options(p): return jsonify({})

# ── Feature builder ───────────────────────────────────────────────────────
def build_features(d):
    orig  = float(d.get('original_price', 0))
    disc  = float(d.get('discounted_price', 0))
    comp  = float(d.get('competitor_price', 0))
    p1m   = float(d.get('price_1_month_ago', orig * 0.95))
    p1w   = float(d.get('price_1_week_ago',  orig * 0.97))
    dpct  = float(d.get('discount_percentage', (orig - disc) / (orig + 1) * 100))
    rat   = float(d.get('rating', 4.0))
    rev   = float(d.get('num_reviews', 100))
    stock = float(d.get('stock_left', 50))
    flash = int(d.get('is_flash_sale', 0))
    cat   = d.get('category', 'Electronics')
    brand = d.get('brand', '')

    try: cat_enc = le_category.transform([cat])[0]
    except: cat_enc = 0
    try: brand_enc = le_brand.transform([brand])[0]
    except: brand_enc = 0

    pim = orig - p1m; piw = orig - p1w
    irm = orig / (p1m + 1); irw = orig / (p1w + 1)
    vsc = disc - comp; cr = disc / (comp + 1)
    ctc = int(disc < comp); sav = orig - disc
    dtc = (orig - comp) / (comp + 1)
    rrs = rat * np.log1p(rev); vi = (rat * rev) / (disc + 1)
    ls  = int(stock < 20); fdc = flash * dpct

    return [p1m, p1w, orig, disc, dpct, comp, rat, rev, stock, flash,
            cat_enc, brand_enc, pim, piw, irm, irw, vsc, cr, ctc,
            sav, dtc, rrs, vi, ls, fdc]

def compute_scores(d, fake_prob):
    orig = float(d.get('original_price', 0))
    disc = float(d.get('discounted_price', 0))
    comp = float(d.get('competitor_price', disc))
    rat  = float(d.get('rating', 4.0))
    rev  = float(d.get('num_reviews', 100))
    price_fairness = float(np.clip(1 - (disc - comp) / (comp + 1), 0, 1))
    quality = float(np.clip((rat-3.7)/(4.9-3.7)*0.6 + np.log1p(rev)/np.log1p(12000)*0.4, 0, 1))
    value = float(np.clip(price_fairness*0.4 + quality*0.4 + (1-fake_prob)*0.2, 0, 1))
    return {
        'price_fairness_score': round(price_fairness, 4),
        'quality_score': round(quality, 4),
        'value_score': round(value, 4),
        'worth_buying': bool((fake_prob < 0.5) and (value >= 0.4))
    }

# ─────────────────────────────────────────────────────────────────────────
# ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({
        'status': 'ok', 'timestamp': datetime.now().isoformat(),
        'model': 'RandomForestClassifier' if model else 'data-only',
        'products': len(products_db), 'version': '2.0'
    })

@app.route('/api/metadata', methods=['GET'])
def get_metadata(): return jsonify(metadata)

@app.route('/api/categories', methods=['GET'])
def get_categories():
    return jsonify({'categories': metadata['categories'], 'brands': metadata['brands']})

@app.route('/api/products', methods=['GET'])
def get_products():
    try:
        page     = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        category = request.args.get('category', '')
        brand    = request.args.get('brand', '')
        search   = request.args.get('search', '').lower()
        fake_only= request.args.get('fake_only', '').lower() == 'true'
        genuine_only = request.args.get('genuine_only', '').lower() == 'true'
        sort_by  = request.args.get('sort_by', 'value_score')
        min_price= float(request.args.get('min_price', 0))
        max_price= float(request.args.get('max_price', 9999999))
        min_rating= float(request.args.get('min_rating', 0))
        flash_only= request.args.get('flash_only', '').lower() == 'true'

        pool = products_db[:]
        if category: pool = [p for p in pool if p.get('category','').lower() == category.lower()]
        if brand:    pool = [p for p in pool if p.get('brand','').lower() == brand.lower()]
        if search:   pool = [p for p in pool if search in p.get('product_name','').lower() or search in p.get('brand','').lower()]
        if fake_only: pool = [p for p in pool if p.get('predicted_fake', 0) == 1]
        if genuine_only: pool = [p for p in pool if p.get('predicted_fake', 0) == 0]
        if flash_only: pool = [p for p in pool if p.get('is_flash_sale')]
        pool = [p for p in pool if min_price <= p.get('discounted_price', 0) <= max_price]
        pool = [p for p in pool if p.get('rating', 0) >= min_rating]

        valid_sorts = {'value_score','rating','discount_percentage','fake_probability','discounted_price','num_reviews'}
        if sort_by in valid_sorts:
            pool.sort(key=lambda p: p.get(sort_by, 0), reverse=(sort_by != 'discounted_price'))
        elif sort_by == 'price_asc':
            pool.sort(key=lambda p: p.get('discounted_price', 0))
        elif sort_by == 'price_desc':
            pool.sort(key=lambda p: p.get('discounted_price', 0), reverse=True)

        total = len(pool)
        start = (page - 1) * per_page
        return jsonify({
            'products': pool[start:start+per_page],
            'total': total, 'page': page, 'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/products/<product_id>', methods=['GET'])
def get_product_detail(product_id):
    """Get single product with full detail + similar products."""
    try:
        product = next((p for p in products_db if p.get('product_id') == product_id), None)
        if not product:
            return jsonify({'error': 'Product not found'}), 404
        
        # Get similar products (same category, not same product)
        similar = [p for p in products_db
                   if p.get('category') == product.get('category')
                   and p.get('product_id') != product_id
                   and p.get('predicted_fake', 1) == 0]
        similar.sort(key=lambda p: p.get('value_score', 0), reverse=True)
        
        return jsonify({
            'product': product,
            'similar_products': similar[:6],
            'price_history': {
                'month_ago': product.get('price_1_month_ago', product.get('original_price')),
                'week_ago':  product.get('price_1_week_ago', product.get('original_price')),
                'current':   product.get('discounted_price')
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/predict', methods=['POST'])
def predict():
    try:
        data = request.get_json(force=True)
        if not data: return jsonify({'error': 'No JSON body'}), 400
        
        if model:
            row  = build_features(data)
            prob = float(model.predict_proba([row])[0][1])
            pred = int(model.predict([row])[0])
        else:
            # Heuristic fallback
            orig = float(data.get('original_price', 1))
            disc = float(data.get('discounted_price', 1))
            comp = float(data.get('competitor_price', orig))
            p1m  = float(data.get('price_1_month_ago', orig * 0.95))
            inflation = (orig - p1m) / (p1m + 1)
            vs_market = (disc - comp) / (comp + 1)
            prob = float(np.clip(0.3 * inflation * 5 + 0.4 * max(vs_market, 0) + 0.3 * (1 if data.get('is_flash_sale') else 0), 0, 1))
            pred = int(prob > 0.5)

        return jsonify({
            'is_fake_discount': pred, 'fake_probability': round(prob, 4),
            'genuine_probability': round(1 - prob, 4),
            'verdict': 'FAKE DISCOUNT' if pred else 'GENUINE DISCOUNT',
            'confidence': round(max(prob, 1-prob)*100, 1),
            'risk_level': 'HIGH' if prob > 0.7 else 'MEDIUM' if prob > 0.4 else 'LOW'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analyze', methods=['POST'])
def analyze():
    try:
        data = request.get_json(force=True)
        if not data: return jsonify({'error': 'No JSON body'}), 400

        if model:
            row  = build_features(data)
            prob = float(model.predict_proba([row])[0][1])
            pred = int(model.predict([row])[0])
        else:
            orig = float(data.get('original_price', 1))
            disc = float(data.get('discounted_price', 1))
            comp = float(data.get('competitor_price', orig))
            p1m  = float(data.get('price_1_month_ago', orig * 0.95))
            inflation = (orig - p1m) / (p1m + 1)
            vs_market = (disc - comp) / (comp + 1)
            prob = float(np.clip(0.3 * inflation * 5 + 0.4 * max(vs_market, 0) + 0.3 * (1 if data.get('is_flash_sale') else 0), 0, 1))
            pred = int(prob > 0.5)

        scores = compute_scores(data, prob)
        orig = float(data.get('original_price', 0))
        disc = float(data.get('discounted_price', 0))
        comp = float(data.get('competitor_price', disc))
        dpct = float(data.get('discount_percentage', (orig-disc)/(orig+1)*100))

        reasons = []
        if prob > 0.5:
            if float(data.get('price_1_month_ago', orig)) < orig * 0.99:
                reasons.append("Price was artificially inflated before discount")
            if disc > comp:
                reasons.append("Discounted price still higher than market average")
            if data.get('is_flash_sale'):
                reasons.append("Flash sale label may create false urgency")
        else:
            reasons = ["Discount appears proportional to historical price",
                       "Price is competitive with market alternatives"]

        return jsonify({
            'prediction': {
                'is_fake_discount': pred, 'fake_probability': round(prob, 4),
                'genuine_probability': round(1-prob, 4),
                'verdict': 'FAKE DISCOUNT' if pred else 'GENUINE DISCOUNT',
                'confidence': round(max(prob, 1-prob)*100, 1),
                'risk_level': 'HIGH' if prob > 0.7 else 'MEDIUM' if prob > 0.4 else 'LOW'
            },
            'scores': scores,
            'price_analysis': {
                'original_price': orig, 'discounted_price': disc,
                'competitor_price': comp, 'actual_savings': round(orig-disc, 2),
                'real_discount_pct': round((orig-disc)/(orig+1)*100, 2),
                'stated_discount_pct': dpct, 'vs_market': round(disc-comp, 2),
                'cheaper_than_market': bool(disc < comp)
            },
            'reasoning': list(dict.fromkeys(reasons))[:3],
            'recommendation': ('✅ Safe to Buy — Genuine discount with good value' if scores['worth_buying'] else '⚠️ Caution — This may not be the deal it appears')
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/recommend', methods=['POST'])
def recommend():
    try:
        data = request.get_json(force=True)
        category = data.get('category', '')
        exclude_id = data.get('exclude_product_id', '')
        top_n = int(data.get('top_n', 8))

        pool = [p for p in products_db
                if p.get('predicted_fake', 1) == 0 and p.get('worth_buying', 0) and
                   p.get('product_id') != exclude_id]
        if category:
            cat_pool = [p for p in pool if p.get('category','').lower() == category.lower()]
            if len(cat_pool) >= 3: pool = cat_pool
        pool.sort(key=lambda p: (p.get('value_score', 0), p.get('rating', 0)), reverse=True)
        return jsonify({'recommendations': pool[:top_n], 'count': len(pool[:top_n]), 'category_filter': category})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/compare', methods=['POST'])
def compare():
    try:
        data = request.get_json(force=True)
        ids = data.get('product_ids', [])
        if len(ids) < 2: return jsonify({'error': 'Provide at least 2 product_ids'}), 400
        id_map = {p['product_id']: p for p in products_db}
        results = [id_map[pid] for pid in ids if pid in id_map]
        if len(results) < 2: return jsonify({'error': 'Products not found'}), 404
        best = max(results, key=lambda p: p.get('value_score', 0))
        return jsonify({'products': results, 'best_value': best['product_id']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stats', methods=['GET'])
def stats():
    try:
        total = len(products_db)
        fake_count = sum(1 for p in products_db if p.get('predicted_fake',0)==1)
        worth = sum(1 for p in products_db if p.get('worth_buying',0)==1)
        by_cat = {}
        for p in products_db:
            cat = p.get('category','Unknown')
            if cat not in by_cat: by_cat[cat] = {'total':0,'fake':0,'genuine':0}
            by_cat[cat]['total'] += 1
            if p.get('predicted_fake',0)==1: by_cat[cat]['fake'] += 1
            else: by_cat[cat]['genuine'] += 1
        return jsonify({
            'total_products': total, 'fake_discounts': fake_count,
            'genuine_discounts': total-fake_count, 'worth_buying': worth,
            'fake_percentage': round(fake_count/total*100, 1) if total else 0,
            'avg_discount_pct': round(sum(p.get('discount_percentage',0) for p in products_db)/total, 2) if total else 0,
            'avg_rating': round(sum(p.get('rating',0) for p in products_db)/total, 2) if total else 0,
            'by_category': by_cat, 'model_metrics': metadata.get('metrics', {}),
            'feature_importances': metadata.get('feature_importances', {})
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/feedback', methods=['POST'])
def feedback():
    try:
        data = request.get_json(force=True)
        entry = {'id': len(feedbacks)+1, 'product_id': data.get('product_id',''),
                 'prediction': data.get('prediction',''), 'user_verdict': data.get('user_verdict',''),
                 'comment': data.get('comment',''), 'helpful': data.get('helpful', True),
                 'timestamp': datetime.now().isoformat()}
        feedbacks.append(entry)
        with open(FEEDBACK_FILE, 'w') as f: json.dump(feedbacks, f, indent=2)
        return jsonify({'message': 'Feedback saved', 'id': entry['id']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    logger.info("🚀 FairDeal API v2.0 starting on http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)
